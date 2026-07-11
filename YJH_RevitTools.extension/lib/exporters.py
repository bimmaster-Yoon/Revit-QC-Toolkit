# -*- coding: utf-8 -*-

import io
import json
import os
import subprocess
import uuid

from System.Diagnostics import Process, ProcessStartInfo
from System.IO import Directory, Path, StreamWriter
from System.Text import UTF8Encoding

from collectors import to_text
from grouping import get_issue_group_fields


def csv_escape(value):
    text = to_text(value)

    if u"," in text or u'"' in text or u"\n" in text or u"\r" in text:
        text = text.replace(u'"', u'""')
        return u'"{0}"'.format(text)

    return text


def get_export_path(export_folder, file_prefix, report_name, timestamp, extension):
    if not export_folder or not Directory.Exists(export_folder):
        raise Exception(
            u"선택한 Export 폴더를 사용할 수 없습니다: {0}".format(
                export_folder
            )
        )

    return Path.Combine(
        export_folder,
        u"{0}_{1}_{2}.{3}".format(
            file_prefix,
            report_name,
            timestamp,
            extension
        )
    )


def write_csv_row(writer, values):
    writer.WriteLine(u",".join([csv_escape(value) for value in values]))


def write_csv_metadata(writer, version, summary_data, qc_status):
    write_csv_row(writer, [u"Report Version", version])
    write_csv_row(writer, [u"QC Status", qc_status])
    write_csv_row(writer, [u"Checked Sheets", summary_data["checked_sheets"]])
    write_csv_row(writer, [u"Checked Views", summary_data["checked_views"]])
    write_csv_row(writer, [u"Sheet Issues", summary_data["sheet_issues"]])
    write_csv_row(writer, [u"View Issues", summary_data["view_issues"]])
    write_csv_row(writer, [u"Parameter Issues", summary_data["parameter_issues"]])
    write_csv_row(writer, [u"Total Issues", summary_data["total_issues"]])
    write_csv_row(writer, [u"High", summary_data["high_count"]])
    write_csv_row(writer, [u"Medium", summary_data["medium_count"]])
    write_csv_row(writer, [u"Low", summary_data["low_count"]])
    writer.WriteLine(u"")


def save_full_csv(
    issue_rows,
    summary_data,
    qc_status,
    timestamp,
    version,
    file_prefix,
    export_folder
):
    csv_path = get_export_path(
        export_folder,
        file_prefix,
        u"Full",
        timestamp,
        u"csv"
    )
    writer = None

    try:
        writer = StreamWriter(csv_path, False, UTF8Encoding(True))
        write_csv_metadata(writer, version, summary_data, qc_status)
        write_csv_row(
            writer,
            [
                u"Category",
                u"Item Type / Number",
                u"Item Name",
                u"Severity",
                u"QC Item",
                u"Issue Message",
                u"Original Issue Detail"
            ]
        )

        for row in issue_rows:
            qc_item, issue_message = get_issue_group_fields(row)
            write_csv_row(
                writer,
                [row[0], row[1], row[2], row[3], qc_item, issue_message, row[4]]
            )
    finally:
        if writer is not None:
            writer.Close()

    return csv_path


def save_summary_csv(
    group_rows,
    summary_data,
    qc_status,
    timestamp,
    version,
    file_prefix,
    export_folder
):
    csv_path = get_export_path(
        export_folder,
        file_prefix,
        u"Summary",
        timestamp,
        u"csv"
    )
    writer = None

    try:
        writer = StreamWriter(csv_path, False, UTF8Encoding(True))
        write_csv_metadata(writer, version, summary_data, qc_status)
        write_csv_row(
            writer,
            [u"Category", u"Item Type", u"QC Item", u"Severity", u"Count", u"Sample Items"]
        )

        for row in group_rows:
            write_csv_row(writer, row)
    finally:
        if writer is not None:
            writer.Close()

    return csv_path


def get_recommendation(category, qc_item, review_message):
    if category == u"Sheet QC":
        if qc_item in [u"Sheet Number", u"Sheet Name"]:
            return u"Complete the required sheet information."
        if qc_item == u"Placed Views":
            return u"Review the sheet composition and place the required views."

    if category == u"View QC":
        if qc_item == u"View Name":
            return u"Rename the view using the project naming standard."
        if qc_item == u"View Scale":
            return u"Set a valid drawing scale for the view."
        if qc_item == u"View Template":
            return u"Apply the approved view template."
        if qc_item == u"Sheet Placement":
            return u"Place the drawing view on a sheet or mark it as working-only."

    if category == u"Parameter QC":
        if review_message == u"Shared Parameter 없음":
            return u"Add the required shared parameter to the category."
        return u"Populate the required parameter value."

    return u"Review the item against the active QC standard."


def extract_element_id(item_name):
    text = to_text(item_name)
    marker = u"[Id:"
    marker_index = text.rfind(marker)

    if marker_index < 0:
        return u""

    end_index = text.find(u"]", marker_index)
    if end_index < 0:
        return u""

    return text[marker_index + len(marker):end_index].strip()


def get_current_value(issue_detail):
    detail = to_text(issue_detail)

    if u"값 비어 있음" in detail:
        return u"(empty)"
    if detail.startswith(u"Shared Parameter 없음:"):
        return u"(parameter missing)"
    if u":" in detail:
        return detail.split(u":", 1)[1].strip()

    return u""


def build_review_group_xlsx_rows(group_rows):
    xlsx_rows = []

    for row in group_rows:
        xlsx_rows.append(
            list(row) + [get_recommendation(row[0], row[2], u"")]
        )

    return xlsx_rows


def build_key_sample_xlsx_rows(key_issue_rows):
    xlsx_rows = []

    for row in key_issue_rows:
        xlsx_rows.append(
            list(row) + [get_recommendation(row[0], row[4], row[5])]
        )

    return xlsx_rows


def build_full_detail_xlsx_rows(issue_rows):
    xlsx_rows = []

    for row in issue_rows:
        qc_item, review_message = get_issue_group_fields(row)
        xlsx_rows.append(
            [
                row[0],
                row[3],
                extract_element_id(row[2]),
                row[1],
                row[2],
                qc_item,
                review_message,
                get_current_value(row[4]),
                get_recommendation(row[0], qc_item, review_message)
            ]
        )

    return xlsx_rows


def get_report_font_name():
    try:
        import clr
        clr.AddReference("System.Drawing")
        from System.Drawing import FontFamily

        available_names = [family.Name.lower() for family in FontFamily.Families]
        if u"suit" in available_names:
            return u"SUIT"
    except Exception:
        pass

    return u"Malgun Gothic"


def apply_xlsx_title(sheet, title, column_count, styles):
    sheet.merge_cells(
        start_row=1,
        start_column=1,
        end_row=1,
        end_column=column_count
    )
    title_cell = sheet.cell(row=1, column=1, value=title)
    title_cell.font = styles["title_font"]
    title_cell.fill = styles["title_fill"]
    title_cell.alignment = styles["title_alignment"]
    sheet.row_dimensions[1].height = 32

    for column_index in range(1, column_count + 1):
        accent_cell = sheet.cell(row=2, column=column_index)
        accent_cell.fill = styles["orange_fill"]
    sheet.row_dimensions[2].height = 4


def apply_xlsx_table(
    sheet,
    title,
    headers,
    rows,
    styles,
    severity_column,
    count_column=None
):
    column_count = len(headers)
    apply_xlsx_title(sheet, title, column_count, styles)

    header_row = 3
    data_start_row = 4

    for column_index, header in enumerate(headers, 1):
        cell = sheet.cell(row=header_row, column=column_index, value=header)
        cell.font = styles["header_font"]
        cell.fill = styles["navy_fill"]
        cell.border = styles["thin_border"]
        cell.alignment = styles["header_alignment"]

    for row_offset, row_values in enumerate(rows):
        row_index = data_start_row + row_offset
        zebra_fill = None

        if row_offset % 2 == 1:
            zebra_fill = styles["zebra_fill"]

        for column_index, value in enumerate(row_values, 1):
            cell = sheet.cell(row=row_index, column=column_index, value=value)
            cell.font = styles["body_font"]
            cell.border = styles["thin_border"]
            cell.alignment = styles["body_alignment"]

            if zebra_fill is not None:
                cell.fill = zebra_fill

            if count_column is not None and column_index == count_column:
                cell.font = styles["body_bold_font"]

        severity_value = to_text(row_values[severity_column - 1])
        severity_cell = sheet.cell(row=row_index, column=severity_column)
        severity_cell.font = styles["body_bold_font"]

        if severity_value == u"High":
            severity_cell.fill = styles["high_fill"]
        elif severity_value == u"Medium":
            severity_cell.fill = styles["medium_fill"]
        elif severity_value == u"Low":
            severity_cell.fill = styles["low_fill"]

    last_row = max(header_row, data_start_row + len(rows) - 1)
    last_column = styles["get_column_letter"](column_count)
    sheet.auto_filter.ref = u"A{0}:{1}{2}".format(
        header_row,
        last_column,
        last_row
    )
    sheet.freeze_panes = u"A4"
    sheet.sheet_view.showGridLines = False
    auto_fit_xlsx_columns(sheet, column_count, last_row)


def auto_fit_xlsx_columns(sheet, column_count, last_row):
    from openpyxl.utils import get_column_letter

    for column_index in range(1, column_count + 1):
        max_length = 0

        for row_index in range(1, last_row + 1):
            value = sheet.cell(row=row_index, column=column_index).value
            if value is None:
                continue
            value_length = len(to_text(value))
            if value_length > max_length:
                max_length = value_length

        adjusted_width = max(12, min(max_length + 3, 52))
        sheet.column_dimensions[get_column_letter(column_index)].width = adjusted_width


def write_xlsx_summary_sheet(
    sheet,
    summary_data,
    qc_status,
    version,
    export_folder,
    report_context,
    styles
):
    apply_xlsx_title(sheet, u"Revit QC Report", 4, styles)
    sheet.sheet_view.showGridLines = False
    result_model = report_context.get("result_model", {})
    kpi = result_model.get("kpi", {})
    issue_counts = result_model.get("issue_count_by_qc", {})

    summary_items = [
        (u"Project", report_context.get("project", u"")),
        (u"Run Mode", report_context.get("run_mode", u"")),
        (u"QC Status", qc_status),
        (
            u"Checked Items",
            kpi.get(
                "checked_items",
                summary_data["checked_sheets"] + summary_data["checked_views"]
            )
        ),
        (
            u"Total Findings",
            kpi.get("total_findings", summary_data["total_issues"])
        ),
        (
            u"Critical Items",
            kpi.get("critical_items", summary_data["high_count"])
        ),
        (
            u"Sheet QC",
            issue_counts.get("sheet_qc", summary_data["sheet_issues"])
        ),
        (
            u"View QC",
            issue_counts.get("view_qc", summary_data["view_issues"])
        ),
        (
            u"Parameter QC",
            issue_counts.get(
                "parameter_qc",
                summary_data["parameter_issues"]
            )
        ),
        (u"Review Groups", report_context.get("review_group_count", 0)),
        (u"High", summary_data["high_count"]),
        (u"Medium", summary_data["medium_count"]),
        (u"Low", summary_data["low_count"]),
        (u"Run Time", report_context.get("run_time", u"")),
        (u"Tool Version", version),
        (u"Export Folder", export_folder)
    ]

    for item_index, item in enumerate(summary_items):
        card_row = 4 + int(item_index / 2)
        card_column = 1 + (item_index % 2) * 2
        label_cell = sheet.cell(row=card_row, column=card_column, value=item[0])
        value_cell = sheet.cell(row=card_row, column=card_column + 1, value=item[1])

        label_cell.font = styles["summary_label_font"]
        label_cell.fill = styles["summary_label_fill"]
        label_cell.border = styles["thin_border"]
        label_cell.alignment = styles["body_alignment"]

        value_cell.font = styles["body_bold_font"]
        value_cell.fill = styles["white_fill"]
        value_cell.border = styles["thin_border"]
        value_cell.alignment = styles["body_alignment"]

        if item[0] in [u"Total Findings", u"Critical Items", u"Review Groups"]:
            value_cell.fill = styles["highlight_fill"]
            value_cell.font = styles["highlight_font"]

    sheet.column_dimensions[u"A"].width = 27
    sheet.column_dimensions[u"B"].width = 34
    sheet.column_dimensions[u"C"].width = 27
    sheet.column_dimensions[u"D"].width = 42

    for row_index in range(4, 12):
        sheet.row_dimensions[row_index].height = 34
    sheet.row_dimensions[11].height = 50

    sheet.freeze_panes = u"A4"


def _export_styled_xlsx_direct(
    issue_rows,
    group_rows,
    summary_data,
    qc_status,
    timestamp,
    version,
    file_prefix,
    export_folder,
    report_context
):
    """openpyxl이 있으면 보고용 XLSX를 생성하고, 없으면 안전하게 skip한다."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except Exception:
        return (
            u"",
            u"Styled XLSX export skipped: openpyxl is not available in the active "
            u"pyRevit engine. Install an IronPython-compatible openpyxl package."
        )

    xlsx_path = get_export_path(
        export_folder,
        file_prefix,
        u"Report",
        timestamp,
        u"xlsx"
    )

    try:
        font_name = get_report_font_name()
        title_color = u"3C4A57"
        navy_color = u"34495A"
        orange_color = u"E97826"
        gray_border_color = u"D9DEE3"

        thin_side = Side(style=u"thin", color=gray_border_color)
        styles = {
            "title_font": Font(
                name=font_name,
                size=16,
                bold=True,
                color=u"FFFFFF"
            ),
            "header_font": Font(
                name=font_name,
                size=10,
                bold=True,
                color=u"FFFFFF"
            ),
            "body_font": Font(name=font_name, size=9, color=u"263645"),
            "body_bold_font": Font(
                name=font_name,
                size=9,
                bold=True,
                color=u"263645"
            ),
            "summary_label_font": Font(
                name=font_name,
                size=9,
                bold=True,
                color=u"263645"
            ),
            "highlight_font": Font(
                name=font_name,
                size=12,
                bold=True,
                color=navy_color
            ),
            "title_fill": PatternFill(
                fill_type=u"solid",
                fgColor=title_color
            ),
            "navy_fill": PatternFill(
                fill_type=u"solid",
                fgColor=navy_color
            ),
            "orange_fill": PatternFill(
                fill_type=u"solid",
                fgColor=orange_color
            ),
            "zebra_fill": PatternFill(
                fill_type=u"solid",
                fgColor=u"F2F4F6"
            ),
            "summary_label_fill": PatternFill(
                fill_type=u"solid",
                fgColor=u"F6F7F8"
            ),
            "white_fill": PatternFill(
                fill_type=u"solid",
                fgColor=u"FFFFFF"
            ),
            "highlight_fill": PatternFill(
                fill_type=u"solid",
                fgColor=u"FFF0E3"
            ),
            "high_fill": PatternFill(
                fill_type=u"solid",
                fgColor=u"FADBD8"
            ),
            "medium_fill": PatternFill(
                fill_type=u"solid",
                fgColor=u"FFF0C2"
            ),
            "low_fill": PatternFill(
                fill_type=u"solid",
                fgColor=u"E9ECEF"
            ),
            "thin_border": Border(
                left=thin_side,
                right=thin_side,
                top=thin_side,
                bottom=thin_side
            ),
            "title_alignment": Alignment(
                horizontal=u"left",
                vertical=u"center"
            ),
            "header_alignment": Alignment(
                horizontal=u"center",
                vertical=u"center",
                wrap_text=True
            ),
            "body_alignment": Alignment(
                vertical=u"top",
                wrap_text=True
            ),
            "get_column_letter": get_column_letter
        }

        workbook = Workbook()
        summary_sheet = workbook.active
        summary_sheet.title = u"SUMMARY"
        write_xlsx_summary_sheet(
            summary_sheet,
            summary_data,
            qc_status,
            version,
            export_folder,
            report_context,
            styles
        )

        review_groups_sheet = workbook.create_sheet(u"Review Groups")
        apply_xlsx_table(
            review_groups_sheet,
            u"Review Groups",
            [
                u"Category",
                u"Item Type",
                u"QC Item",
                u"Severity",
                u"Count",
                u"Sample Items",
                u"Recommendation"
            ],
            build_review_group_xlsx_rows(group_rows),
            styles,
            4,
            5
        )

        key_samples_sheet = workbook.create_sheet(u"Key Samples")
        apply_xlsx_table(
            key_samples_sheet,
            u"Key Samples",
            [
                u"Category",
                u"Item Type",
                u"Item Name",
                u"Severity",
                u"QC Item",
                u"Review Message",
                u"Recommendation"
            ],
            build_key_sample_xlsx_rows(
                report_context.get("key_issue_rows", [])
            ),
            styles,
            4
        )

        full_detail_sheet = workbook.create_sheet(u"Full Detail")
        apply_xlsx_table(
            full_detail_sheet,
            u"Full Detail",
            [
                u"Category",
                u"Severity",
                u"ElementId",
                u"ElementType",
                u"Name",
                u"QC Item",
                u"Review Message",
                u"Current Value",
                u"Recommendation"
            ],
            build_full_detail_xlsx_rows(issue_rows),
            styles,
            2
        )

        workbook.save(xlsx_path)
        return xlsx_path, u""
    except Exception as ex:
        return u"", u"Styled XLSX export failed: {0}".format(to_text(ex))


def build_styled_xlsx_payload(
    issue_rows,
    group_rows,
    summary_data,
    qc_status,
    version,
    export_folder,
    report_context
):
    active_config_path = to_text(report_context.get("active_config", u""))
    active_config_display = to_text(
        report_context.get("active_config_display", u"")
    )
    if not active_config_display:
        active_config_display = os.path.basename(active_config_path)
    export_folder_path = to_text(export_folder)

    payload = {
        "result_model": report_context.get("result_model", {}),
        "summary_data": summary_data,
        "review_groups": build_review_group_xlsx_rows(group_rows),
        "key_samples": build_key_sample_xlsx_rows(
            report_context.get("key_issue_rows", [])
        ),
        "full_detail": build_full_detail_xlsx_rows(issue_rows),
        "metadata": {
            "project": to_text(report_context.get("project", u"")),
            "active_config": active_config_display,
            "active_preset": to_text(
                report_context.get("active_preset", u"")
            ),
            "run_mode": to_text(report_context.get("run_mode", u"")),
            "qc_status": to_text(qc_status),
            "checked_parameter_elements": report_context.get(
                "checked_parameter_elements",
                0
            ),
            "review_group_count": report_context.get("review_group_count", 0),
            "run_time": to_text(report_context.get("run_time", u"")),
            "export_time": to_text(report_context.get("export_time", u"")),
            "tool_version": to_text(version),
            "export_folder": os.path.basename(
                os.path.normpath(export_folder_path)
            )
        },
        "tool_version": to_text(version),
        "active_config": active_config_display,
        "export_folder": export_folder_path
    }

    return payload


def quote_process_arguments(arguments):
    return subprocess.list2cmdline([to_text(value) for value in arguments])


def format_process_command(executable_path, arguments):
    return u"FileName={0} | Arguments={1}".format(
        to_text(executable_path),
        quote_process_arguments(arguments)
    )


def run_external_process(
    executable_path,
    arguments,
    timeout_milliseconds=30000,
    working_directory=u""
):
    try:
        start_info = ProcessStartInfo()
        start_info.FileName = executable_path
        start_info.Arguments = quote_process_arguments(arguments)
        start_info.UseShellExecute = False
        start_info.CreateNoWindow = True
        start_info.RedirectStandardOutput = True
        start_info.RedirectStandardError = True
        start_info.EnvironmentVariables["PYTHONIOENCODING"] = "utf-8"

        if working_directory:
            start_info.WorkingDirectory = working_directory

        process = Process()
        process.StartInfo = start_info
        process.Start()

        if not process.WaitForExit(timeout_milliseconds):
            try:
                process.Kill()
            except Exception:
                pass
            return -1, u"", u"External Python process timed out."

        standard_output = to_text(process.StandardOutput.ReadToEnd()).strip()
        standard_error = to_text(process.StandardError.ReadToEnd()).strip()
        return process.ExitCode, standard_output, standard_error
    except Exception as ex:
        return -1, u"", to_text(ex)


def get_external_python_candidates(configured_python_path):
    candidates = []
    configured_path = to_text(configured_python_path).strip()

    if (
        len(configured_path) >= 2
        and configured_path[0] == u'"'
        and configured_path[-1] == u'"'
    ):
        configured_path = configured_path[1:-1]

    configured_path = os.path.expandvars(configured_path)

    if configured_path:
        candidates.append(
            {
                "name": u"Configured external Python",
                "executable": configured_path,
                "prefix_arguments": []
            }
        )

    candidates.extend(
        [
            {
                "name": u"py -3",
                "executable": u"py",
                "prefix_arguments": [u"-3"]
            },
            {
                "name": u"python",
                "executable": u"python",
                "prefix_arguments": []
            },
            {
                "name": u"python3",
                "executable": u"python3",
                "prefix_arguments": []
            }
        ]
    )

    return candidates


def find_external_python_with_openpyxl(
    configured_python_path,
    debug_lines=None
):
    probe_code = (
        u"import sys, openpyxl; "
        u"sys.stdout.write(sys.executable + '|' + openpyxl.__version__)"
    )
    probe_errors = []

    for candidate in get_external_python_candidates(configured_python_path):
        probe_arguments = list(candidate["prefix_arguments"])
        probe_arguments.extend([u"-c", probe_code])

        if debug_lines is not None:
            debug_lines.append(u"python_candidate: {0}".format(candidate["name"]))
            debug_lines.append(
                u"python_probe_command: {0}".format(
                    format_process_command(
                        candidate["executable"],
                        probe_arguments
                    )
                )
            )

        exit_code, standard_output, standard_error = run_external_process(
            candidate["executable"],
            probe_arguments,
            15000
        )

        if debug_lines is not None:
            debug_lines.append(u"python_probe_exit_code: {0}".format(exit_code))
            debug_lines.append(
                u"python_probe_stdout: {0}".format(standard_output or u"(empty)")
            )
            debug_lines.append(
                u"python_probe_stderr: {0}".format(standard_error or u"(empty)")
            )

        if exit_code == 0 and standard_output:
            candidate["probe_output"] = standard_output
            return candidate, u""

        error_text = standard_error or standard_output or u"not available"
        probe_errors.append(
            u"{0}: {1}".format(candidate["name"], error_text)
        )

    return None, u"; ".join(probe_errors)


def find_external_python_with_reportlab(
    configured_python_path,
    debug_lines=None
):
    probe_code = (
        u"import sys, reportlab; "
        u"sys.stdout.write(sys.executable + '|' + reportlab.Version)"
    )
    probe_errors = []
    for candidate in get_external_python_candidates(configured_python_path):
        probe_arguments = list(candidate["prefix_arguments"])
        probe_arguments.extend([u"-c", probe_code])
        exit_code, standard_output, standard_error = run_external_process(
            candidate["executable"], probe_arguments, 15000
        )
        if debug_lines is not None:
            debug_lines.append(
                u"pdf_python_candidate: {0} | exit={1}".format(
                    candidate["name"],
                    exit_code
                )
            )
        if exit_code == 0 and standard_output:
            candidate["probe_output"] = standard_output
            return candidate, u""
        probe_errors.append(
            u"{0}: {1}".format(
                candidate["name"],
                standard_error or standard_output or u"not available"
            )
        )
    return None, u"; ".join(probe_errors)


def probe_external_python_path(executable_path):
    """선택된 Python 한 개만 실행하고 openpyxl 환경을 진단한다."""
    python_path = to_text(executable_path).strip()
    status = {
        "external_python_path": python_path,
        "external_python_detected": u"No",
        "openpyxl_available": u"No",
        "python_detail": u"",
        "openpyxl_version": u"",
        "probe_error": u""
    }

    if (
        len(python_path) >= 2
        and python_path[0] == u'"'
        and python_path[-1] == u'"'
    ):
        python_path = python_path[1:-1]

    python_path = os.path.expandvars(python_path)
    status["external_python_path"] = python_path

    if not python_path:
        status["probe_error"] = u"External Python path is empty."
        return status

    version_probe = (
        u"import sys; "
        u"sys.stdout.write(sys.executable + '|' + sys.version.split()[0])"
    )
    exit_code, standard_output, standard_error = run_external_process(
        python_path,
        [u"-c", version_probe],
        15000
    )

    if exit_code != 0:
        status["probe_error"] = (
            standard_error
            or standard_output
            or u"Selected Python could not be executed."
        )
        return status

    status["external_python_detected"] = u"Yes"
    status["python_detail"] = standard_output
    openpyxl_probe = (
        u"import openpyxl; "
        u"import sys; sys.stdout.write(openpyxl.__version__)"
    )
    exit_code, standard_output, standard_error = run_external_process(
        python_path,
        [u"-c", openpyxl_probe],
        15000
    )

    if exit_code == 0 and standard_output:
        status["openpyxl_available"] = u"Yes"
        status["openpyxl_version"] = standard_output
        return status

    status["probe_error"] = (
        standard_error
        or standard_output
        or u"Selected Python does not have openpyxl."
    )
    return status


def get_xlsx_debug_log_path(reports_dir):
    return os.path.join(reports_dir, "xlsx_helper_debug.log")


def write_xlsx_debug_log(debug_log_path, debug_lines):
    try:
        debug_folder = os.path.dirname(debug_log_path)
        if not os.path.isdir(debug_folder):
            os.makedirs(debug_folder)

        with io.open(debug_log_path, "w", encoding="utf-8") as debug_file:
            debug_file.write(u"\n".join([to_text(line) for line in debug_lines]))
        return u""
    except Exception as ex:
        return to_text(ex)


def finalize_xlsx_result(
    debug_log_path,
    debug_lines,
    saved_xlsx_path,
    error_message
):
    debug_lines.append(
        u"xlsx_exists: {0}".format(
            bool(saved_xlsx_path and os.path.isfile(saved_xlsx_path))
        )
    )
    debug_lines.append(
        u"failure_reason: {0}".format(error_message or u"(none)")
    )
    log_error = write_xlsx_debug_log(debug_log_path, debug_lines)

    if error_message:
        result_message = u"{0} Debug log: {1}".format(
            error_message,
            debug_log_path
        )
        if log_error:
            result_message = u"{0} (debug log write failed: {1})".format(
                result_message,
                log_error
            )
        return u"", result_message

    return saved_xlsx_path, u""


def get_xlsx_environment_status(
    configured_python_path,
    extension_dir,
    reports_dir
):
    probe_lines = []
    candidate, probe_error = find_external_python_with_openpyxl(
        configured_python_path,
        probe_lines
    )
    helper_path = os.path.join(extension_dir, "tools", "make_styled_xlsx.py")
    status = {
        "external_python_detected": u"No",
        "openpyxl_available": u"No",
        "python_detail": u"",
        "probe_error": probe_error,
        "helper_path": helper_path,
        "helper_exists": os.path.isfile(helper_path),
        "debug_log_path": get_xlsx_debug_log_path(reports_dir)
    }

    if candidate is not None:
        status["external_python_detected"] = u"Yes"
        status["openpyxl_available"] = u"Yes"
        status["python_detail"] = candidate.get("probe_output", u"")

    return status


def export_styled_xlsx(
    issue_rows,
    group_rows,
    summary_data,
    qc_status,
    timestamp,
    version,
    file_prefix,
    export_folder,
    report_context
):
    """IronPython에서 JSON을 만들고 외부 Python helper로 XLSX를 생성한다."""
    extension_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    reports_dir = to_text(report_context.get("reports_dir", u""))

    if not reports_dir:
        reports_dir = os.path.join(extension_dir, "reports")

    debug_log_path = get_xlsx_debug_log_path(reports_dir)
    helper_path = os.path.join(extension_dir, "tools", "make_styled_xlsx.py")
    xlsx_path = get_export_path(
        export_folder,
        file_prefix,
        u"Report",
        timestamp,
        u"xlsx"
    )
    debug_lines = [
        u"styled_xlsx_selected: True",
        u"export_folder: {0}".format(export_folder),
        u"output_xlsx_path: {0}".format(xlsx_path),
        u"helper_script_path: {0}".format(helper_path),
        u"helper_script_exists: {0}".format(os.path.isfile(helper_path))
    ]

    configured_python_path = report_context.get("external_python_path", u"")
    debug_lines.append(
        u"configured_external_python_path: {0}".format(
            configured_python_path or u"(empty)"
        )
    )
    python_candidate, probe_error = find_external_python_with_openpyxl(
        configured_python_path,
        debug_lines
    )

    if python_candidate is None:
        return finalize_xlsx_result(
            debug_log_path,
            debug_lines,
            u"",
            u"Styled XLSX export requires external Python with openpyxl. "
            u"Install: py -3 -m pip install openpyxl. Probe: {0}".format(
                probe_error
            )
        )

    debug_lines.append(
        u"external_python_detected: {0}".format(
            python_candidate.get("probe_output", u"")
        )
    )

    if not os.path.isfile(helper_path):
        return finalize_xlsx_result(
            debug_log_path,
            debug_lines,
            u"",
            u"Styled XLSX helper not found: {0}".format(helper_path)
        )

    temp_dir = os.path.join(reports_dir, "temp")
    if not os.path.isdir(temp_dir):
        os.makedirs(temp_dir)

    json_path = os.path.join(
        temp_dir,
        u"styled_xlsx_{0}_{1}.json".format(timestamp, uuid.uuid4().hex)
    )
    debug_lines.append(u"temp_json_path: {0}".format(json_path))
    keep_temp_json = bool(report_context.get("debug_keep_temp_json", False))
    saved_xlsx_path = u""
    error_message = u""

    try:
        payload = build_styled_xlsx_payload(
            issue_rows,
            group_rows,
            summary_data,
            qc_status,
            version,
            export_folder,
            report_context
        )
        json_text = to_text(json.dumps(payload, ensure_ascii=True))

        with io.open(json_path, "w", encoding="utf-8") as json_file:
            json_file.write(json_text)

        json_exists = os.path.isfile(json_path)
        debug_lines.append(u"temp_json_exists: {0}".format(json_exists))
        if not json_exists:
            error_message = u"Styled XLSX temp JSON was not created."
        else:
            helper_arguments = list(python_candidate["prefix_arguments"])
            helper_arguments.extend([helper_path, json_path, xlsx_path])
            debug_lines.append(
                u"external_python_command: {0}".format(
                    format_process_command(
                        python_candidate["executable"],
                        helper_arguments
                    )
                )
            )
            exit_code, standard_output, standard_error = run_external_process(
                python_candidate["executable"],
                helper_arguments,
                120000,
                os.path.dirname(helper_path)
            )
            debug_lines.append(
                u"external_python_stdout: {0}".format(
                    standard_output or u"(empty)"
                )
            )
            debug_lines.append(
                u"external_python_stderr: {0}".format(
                    standard_error or u"(empty)"
                )
            )
            debug_lines.append(
                u"external_python_exit_code: {0}".format(exit_code)
            )
            xlsx_exists = os.path.isfile(xlsx_path)
            debug_lines.append(u"xlsx_exists_after_helper: {0}".format(xlsx_exists))

            if exit_code != 0:
                helper_error = standard_error or standard_output or u"unknown error"
                error_message = u"Styled XLSX helper failed: {0}".format(
                    helper_error
                )
            elif not xlsx_exists:
                error_message = (
                    u"Styled XLSX helper completed without an output file."
                )
            else:
                saved_xlsx_path = xlsx_path
    except Exception as ex:
        error_message = u"Styled XLSX export failed: {0}".format(to_text(ex))
    finally:
        if not keep_temp_json:
            try:
                if os.path.isfile(json_path):
                    os.remove(json_path)
                debug_lines.append(
                    u"temp_json_deleted: {0}".format(
                        not os.path.isfile(json_path)
                    )
                )
            except Exception:
                debug_lines.append(u"temp_json_deleted: False")
        else:
            debug_lines.append(u"temp_json_deleted: False (debug keep enabled)")

    return finalize_xlsx_result(
        debug_log_path,
        debug_lines,
        saved_xlsx_path,
        error_message
    )


def export_compact_summary_pdf(
    result_model,
    timestamp,
    file_prefix,
    export_folder,
    report_context
):
    """공통 result model을 외부 Python helper로 1-page PDF로 출력한다."""
    extension_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    reports_dir = to_text(report_context.get("reports_dir", u""))
    if not reports_dir:
        reports_dir = os.path.join(extension_dir, "reports")
    helper_path = os.path.join(
        extension_dir,
        "tools",
        "make_compact_summary_pdf.py"
    )
    pdf_path = get_export_path(
        export_folder,
        file_prefix,
        u"Compact_Summary",
        timestamp,
        u"pdf"
    )
    if not os.path.isfile(helper_path):
        return u"", u"Compact Summary PDF helper not found: {0}".format(helper_path)
    configured_python_path = report_context.get("external_python_path", u"")
    python_candidate, probe_error = find_external_python_with_reportlab(
        configured_python_path
    )
    if python_candidate is None:
        return (
            u"",
            u"Compact Summary PDF requires external Python with reportlab. "
            u"Install reportlab in the configured Python. Probe: {0}".format(
                probe_error
            )
        )
    temp_dir = os.path.join(reports_dir, "temp")
    if not os.path.isdir(temp_dir):
        os.makedirs(temp_dir)
    json_path = os.path.join(
        temp_dir,
        u"compact_summary_{0}_{1}.json".format(
            timestamp,
            uuid.uuid4().hex
        )
    )
    try:
        with io.open(json_path, "w", encoding="utf-8") as json_file:
            json_file.write(to_text(json.dumps(result_model, ensure_ascii=True)))
        helper_arguments = list(python_candidate["prefix_arguments"])
        helper_arguments.extend([helper_path, json_path, pdf_path])
        exit_code, standard_output, standard_error = run_external_process(
            python_candidate["executable"],
            helper_arguments,
            120000,
            os.path.dirname(helper_path)
        )
        if exit_code != 0:
            return (
                u"",
                u"Compact Summary PDF export failed: {0}".format(
                    standard_error or standard_output or u"unknown error"
                )
            )
        if not os.path.isfile(pdf_path):
            return u"", u"PDF helper completed without an output file."
        return pdf_path, u""
    except Exception as ex:
        return u"", u"Compact Summary PDF export failed: {0}".format(to_text(ex))
    finally:
        try:
            if os.path.isfile(json_path):
                os.remove(json_path)
        except Exception:
            pass
