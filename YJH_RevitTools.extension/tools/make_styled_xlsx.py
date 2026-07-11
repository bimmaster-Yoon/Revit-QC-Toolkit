# -*- coding: utf-8 -*-

from __future__ import print_function

import json
import os
import sys

from openpyxl import Workbook
from openpyxl.worksheet.page import PageMargins
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


HEADER_NAVY = "536777"
TITLE_NAVY = "4A5B6A"
TEXT_NAVY = "263645"
SOFT_NAVY_LINE = "D6DDE3"
LIGHT_BORDER = "E1E5E8"
LIGHT_FILL = "F4F6F8"
ZEBRA_FILL = "F4F6F8"
WARNING_FILL = "FFF3E8"
MEDIUM_FILL = "FFF7D6"
LOW_FILL = "F1F3F5"


def load_payload(json_path):
    with open(json_path, "r", encoding="utf-8") as input_file:
        return json.load(input_file)


def has_suit_font():
    try:
        import winreg

        font_key_path = r"SOFTWARE\Microsoft\Windows NT\CurrentVersion\Fonts"
        font_key = winreg.OpenKey(winreg.HKEY_LOCAL_MACHINE, font_key_path)
        value_index = 0

        while True:
            try:
                value_name, value_data, value_type = winreg.EnumValue(
                    font_key,
                    value_index
                )
                combined_text = "{0} {1}".format(value_name, value_data).lower()
                if "suit" in combined_text:
                    winreg.CloseKey(font_key)
                    return True
                value_index += 1
            except OSError:
                break

        winreg.CloseKey(font_key)
    except Exception:
        pass

    return False


def get_report_font_name():
    if has_suit_font():
        return "SUIT"
    return "Malgun Gothic"


def build_styles():
    font_name = get_report_font_name()
    thin_side = Side(style="thin", color=LIGHT_BORDER)
    soft_navy_side = Side(style="thin", color=SOFT_NAVY_LINE)

    return {
        "title_font": Font(
            name=font_name,
            size=16,
            bold=True,
            color="FFFFFF"
        ),
        "header_font": Font(
            name=font_name,
            size=10,
            bold=True,
            color="FFFFFF"
        ),
        "summary_title_font": Font(
            name=font_name,
            size=18,
            bold=True,
            color="FFFFFF"
        ),
        "body_font": Font(name=font_name, size=10, color=TEXT_NAVY),
        "body_bold_font": Font(
            name=font_name,
            size=10,
            bold=True,
            color=TEXT_NAVY
        ),
        "table_body_font": Font(name=font_name, size=9.5, color=TEXT_NAVY),
        "table_body_bold_font": Font(
            name=font_name,
            size=9.5,
            bold=True,
            color=TEXT_NAVY
        ),
        "detail_body_font": Font(name=font_name, size=9, color=TEXT_NAVY),
        "detail_body_bold_font": Font(
            name=font_name,
            size=9,
            bold=True,
            color=TEXT_NAVY
        ),
        "summary_label_font": Font(
            name=font_name,
            size=9,
            bold=True,
            color=TEXT_NAVY
        ),
        "metadata_value_font": Font(
            name=font_name,
            size=9,
            color=TEXT_NAVY
        ),
        "kpi_value_font": Font(
            name=font_name,
            size=16,
            bold=True,
            color=TEXT_NAVY
        ),
        "section_font": Font(
            name=font_name,
            size=10,
            bold=True,
            color=TEXT_NAVY
        ),
        "subtitle_font": Font(name=font_name, size=9, color="6C757D"),
        "title_fill": PatternFill(fill_type="solid", fgColor=TITLE_NAVY),
        "navy_fill": PatternFill(fill_type="solid", fgColor=HEADER_NAVY),
        "soft_line_fill": PatternFill(
            fill_type="solid",
            fgColor=SOFT_NAVY_LINE
        ),
        "zebra_fill": PatternFill(fill_type="solid", fgColor=ZEBRA_FILL),
        "summary_label_fill": PatternFill(
            fill_type="solid",
            fgColor=LIGHT_FILL
        ),
        "white_fill": PatternFill(fill_type="solid", fgColor="FFFFFF"),
        "warning_fill": PatternFill(fill_type="solid", fgColor=WARNING_FILL),
        "high_fill": PatternFill(fill_type="solid", fgColor=WARNING_FILL),
        "medium_fill": PatternFill(fill_type="solid", fgColor=MEDIUM_FILL),
        "low_fill": PatternFill(fill_type="solid", fgColor=LOW_FILL),
        "thin_border": Border(
            left=thin_side,
            right=thin_side,
            top=thin_side,
            bottom=thin_side
        ),
        "header_border": Border(
            left=soft_navy_side,
            right=soft_navy_side,
            top=soft_navy_side,
            bottom=soft_navy_side
        ),
        "title_border": Border(bottom=soft_navy_side),
        "title_alignment": Alignment(
            horizontal="left",
            vertical="center",
            indent=1
        ),
        "header_alignment": Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        ),
        "body_alignment": Alignment(
            horizontal="left",
            vertical="top",
            wrap_text=True
        ),
        "center_alignment": Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        ),
        "right_alignment": Alignment(
            horizontal="right",
            vertical="center",
            wrap_text=True
        )
    }


def apply_title(sheet, title, column_count, styles):
    sheet.merge_cells(
        start_row=1,
        start_column=1,
        end_row=1,
        end_column=column_count
    )
    title_cell = sheet.cell(row=1, column=1, value=title)
    title_cell.font = styles["title_font"]
    title_cell.fill = styles["title_fill"]
    title_cell.alignment = styles["title_alignment"]
    sheet.row_dimensions[1].height = 34

    for column_index in range(1, column_count + 1):
        accent_cell = sheet.cell(row=2, column=column_index)
        accent_cell.fill = styles["soft_line_fill"]
    sheet.row_dimensions[2].height = 2


def apply_common_sheet_settings(sheet, print_title_rows, print_area):
    sheet.sheet_view.showGridLines = False
    sheet.sheet_view.zoomScale = 90
    sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_margins = PageMargins(
        left=0.25,
        right=0.25,
        top=0.35,
        bottom=0.35,
        header=0.15,
        footer=0.15
    )
    sheet.print_title_rows = print_title_rows
    sheet.print_area = print_area


def apply_table(
    sheet,
    title,
    headers,
    rows,
    styles,
    severity_column,
    column_widths,
    count_column=None,
    numeric_columns=None,
    center_columns=None,
    detail_mode=False
):
    column_count = len(headers)
    apply_title(sheet, title, column_count, styles)
    header_row = 3
    data_start_row = 4

    for column_index, header in enumerate(headers, 1):
        cell = sheet.cell(row=header_row, column=column_index, value=header)
        cell.font = styles["header_font"]
        cell.fill = styles["navy_fill"]
        cell.border = styles["header_border"]
        cell.alignment = styles["header_alignment"]
    sheet.row_dimensions[header_row].height = 28

    body_font = styles["detail_body_font"] if detail_mode else styles["table_body_font"]
    bold_font = (
        styles["detail_body_bold_font"]
        if detail_mode
        else styles["table_body_bold_font"]
    )
    numeric_columns = numeric_columns or []
    center_columns = center_columns or []

    for row_offset, row_values in enumerate(rows):
        row_index = data_start_row + row_offset
        zebra_fill = styles["zebra_fill"] if row_offset % 2 == 1 else None

        for column_index, value in enumerate(row_values, 1):
            cell = sheet.cell(row=row_index, column=column_index, value=value)
            cell.font = body_font
            cell.border = styles["thin_border"]
            cell.alignment = styles["body_alignment"]

            if zebra_fill is not None:
                cell.fill = zebra_fill
            if count_column is not None and column_index == count_column:
                cell.font = bold_font
            if column_index in numeric_columns:
                cell.alignment = styles["right_alignment"]
            elif column_index in center_columns:
                cell.alignment = styles["center_alignment"]

        severity_value = str(row_values[severity_column - 1])
        severity_cell = sheet.cell(row=row_index, column=severity_column)
        severity_cell.font = bold_font
        severity_cell.alignment = styles["center_alignment"]

        if severity_value == "High":
            severity_cell.fill = styles["high_fill"]
        elif severity_value == "Medium":
            severity_cell.fill = styles["medium_fill"]
        elif severity_value == "Low":
            severity_cell.fill = styles["low_fill"]
        sheet.row_dimensions[row_index].height = 30

    last_row = max(header_row, data_start_row + len(rows) - 1)
    last_column = get_column_letter(column_count)
    sheet.auto_filter.ref = "A{0}:{1}{2}".format(
        header_row,
        last_column,
        last_row
    )
    sheet.freeze_panes = "A4"
    for column_index, width in enumerate(column_widths, 1):
        sheet.column_dimensions[chr(64 + column_index)].width = width
    apply_common_sheet_settings(
        sheet,
        "$1:$3",
        "A1:{0}{1}".format(chr(64 + column_count), last_row)
    )


def get_display_config_name(config_value):
    file_name = os.path.basename(str(config_value or "").strip())
    if not file_name:
        return ""
    if "(" in file_name and file_name.endswith(")"):
        return file_name
    if file_name.lower() == "qc_config_default.json":
        return "Default QC Config ({0})".format(file_name)

    display_name = os.path.splitext(file_name)[0].replace("_", " ").strip()
    return "{0} ({1})".format(display_name.title(), file_name)


def get_short_folder_name(folder_value):
    folder_text = str(folder_value or "").strip()
    if not folder_text:
        return ""

    normalized_path = os.path.normpath(folder_text)
    short_name = os.path.basename(normalized_path)
    return short_name or folder_text


def apply_block_style(
    sheet,
    min_row,
    max_row,
    min_column,
    max_column,
    font,
    fill,
    border,
    alignment
):
    for row_index in range(min_row, max_row + 1):
        for column_index in range(min_column, max_column + 1):
            cell = sheet.cell(row=row_index, column=column_index)
            cell.font = font
            cell.fill = fill
            cell.border = border
            cell.alignment = alignment


def write_summary_sheet(sheet, payload, styles):
    summary_data = payload.get("summary_data", {})
    metadata = payload.get("metadata", {})
    result_model = payload.get("result_model", {})
    kpi = result_model.get("kpi", {})
    issue_counts = result_model.get("issue_count_by_qc", {})
    generated_at = metadata.get("export_time", "") or metadata.get(
        "run_time",
        ""
    )

    sheet.merge_cells("A1:H2")
    title_cell = sheet["A1"]
    title_cell.value = "Revit QC Report Summary"
    title_cell.font = styles["summary_title_font"]
    title_cell.fill = styles["title_fill"]
    title_cell.border = styles["title_border"]
    title_cell.alignment = styles["title_alignment"]
    sheet.row_dimensions[1].height = 18
    sheet.row_dimensions[2].height = 14

    sheet.merge_cells("A3:H3")
    subtitle = "Project: {0}  |  Run Mode: {1}  |  Generated: {2}".format(
        metadata.get("project", ""),
        metadata.get("run_mode", ""),
        generated_at
    )
    sheet["A3"] = subtitle
    sheet["A3"].font = styles["subtitle_font"]
    sheet["A3"].alignment = styles["body_alignment"]
    sheet.row_dimensions[3].height = 20

    for column_index in range(1, 9):
        sheet.cell(row=4, column=column_index).fill = styles["soft_line_fill"]
    sheet.row_dimensions[4].height = 2

    kpi_items = [
        ("Checked Items", kpi.get("checked_items", summary_data.get("checked_sheets", 0) + summary_data.get("checked_views", 0)), 1, 2, False),
        ("Total Findings", kpi.get("total_findings", summary_data.get("total_issues", 0)), 3, 4, True),
        ("Critical Items", kpi.get("critical_items", summary_data.get("high_count", 0)), 5, 6, True),
        ("Review Groups", result_model.get("review_group_count", metadata.get("review_group_count", 0)), 7, 8, False)
    ]
    for label, value, start_column, end_column, is_warning in kpi_items:
        if start_column != end_column:
            sheet.merge_cells(
                start_row=5,
                start_column=start_column,
                end_row=5,
                end_column=end_column
            )
        sheet.merge_cells(
            start_row=6,
            start_column=start_column,
            end_row=7,
            end_column=end_column
        )
        card_fill = (
            styles["warning_fill"]
            if is_warning
            else styles["summary_label_fill"]
        )
        apply_block_style(
            sheet,
            5,
            7,
            start_column,
            end_column,
            styles["summary_label_font"],
            card_fill,
            styles["thin_border"],
            styles["center_alignment"]
        )
        label_cell = sheet.cell(row=5, column=start_column, value=label)
        value_cell = sheet.cell(row=6, column=start_column, value=value)
        value_cell.font = styles["kpi_value_font"]
        value_cell.number_format = "#,##0"

    sheet.merge_cells("A9:H9")
    sheet["A9"] = "STATUS & ISSUE COUNT BY QC"
    sheet["A9"].font = styles["section_font"]
    sheet["A9"].fill = styles["summary_label_fill"]
    sheet["A9"].border = styles["thin_border"]
    sheet["A9"].alignment = styles["body_alignment"]

    status_items = [
        (
            "QC Status",
            result_model.get("qc_status", metadata.get("qc_status", "")),
            "warning_fill",
            False,
            1,
            2
        ),
        (
            "Sheet QC",
            issue_counts.get("sheet_qc", summary_data.get("sheet_issues", 0)),
            "white_fill",
            True,
            3,
            4
        ),
        (
            "View QC",
            issue_counts.get("view_qc", summary_data.get("view_issues", 0)),
            "white_fill",
            True,
            5,
            6
        ),
        (
            "Parameter QC",
            issue_counts.get("parameter_qc", summary_data.get("parameter_issues", 0)),
            "white_fill",
            True,
            7,
            8
        ),
    ]
    for item in status_items:
        start_column = item[4]
        end_column = item[5]
        if start_column != end_column:
            sheet.merge_cells(
                start_row=10,
                start_column=start_column,
                end_row=10,
                end_column=end_column
            )
            sheet.merge_cells(
                start_row=11,
                start_column=start_column,
                end_row=11,
                end_column=end_column
            )
        apply_block_style(
            sheet,
            10,
            10,
            start_column,
            end_column,
            styles["summary_label_font"],
            styles["summary_label_fill"],
            styles["thin_border"],
            styles["center_alignment"]
        )
        apply_block_style(
            sheet,
            11,
            11,
            start_column,
            end_column,
            styles["body_bold_font"],
            styles[item[2]],
            styles["thin_border"],
            styles["center_alignment"]
        )
        label_cell = sheet.cell(row=10, column=start_column, value=item[0])
        value_cell = sheet.cell(row=11, column=start_column, value=item[1])
        label_cell.font = styles["summary_label_font"]
        value_cell.font = styles["body_bold_font"]
        if item[3]:
            value_cell.number_format = "#,##0"

    sheet.merge_cells("A14:H14")
    sheet["A14"] = "REPORT METADATA"
    sheet["A14"].font = styles["section_font"]
    sheet["A14"].fill = styles["summary_label_fill"]
    sheet["A14"].border = styles["thin_border"]
    sheet["A14"].alignment = styles["body_alignment"]

    metadata_layout = [
        (15, 1, 2, "Project", 3, 4, metadata.get("project", "")),
        (15, 5, 6, "Run Mode", 7, 8, metadata.get("run_mode", "")),
        (16, 1, 2, "Generated At", 3, 4, generated_at),
        (16, 5, 6, "Tool Version", 7, 8, metadata.get("tool_version", "")),
        (
            17,
            1,
            2,
            "Active Config",
            3,
            8,
            get_display_config_name(metadata.get("active_config", ""))
        ),
        (
            18,
            1,
            2,
            "Export Folder",
            3,
            8,
            get_short_folder_name(metadata.get("export_folder", ""))
        )
    ]
    for (
        row_index,
        label_start,
        label_end,
        label,
        value_start,
        value_end,
        value
    ) in metadata_layout:
        if label_start != label_end:
            sheet.merge_cells(
                start_row=row_index,
                start_column=label_start,
                end_row=row_index,
                end_column=label_end
            )
        if value_start != value_end:
            sheet.merge_cells(
                start_row=row_index,
                start_column=value_start,
                end_row=row_index,
                end_column=value_end
            )
        apply_block_style(
            sheet,
            row_index,
            row_index,
            label_start,
            label_end,
            styles["summary_label_font"],
            styles["summary_label_fill"],
            styles["thin_border"],
            styles["body_alignment"]
        )
        apply_block_style(
            sheet,
            row_index,
            row_index,
            value_start,
            value_end,
            styles["metadata_value_font"],
            styles["white_fill"],
            styles["thin_border"],
            styles["body_alignment"]
        )
        sheet.cell(row=row_index, column=label_start, value=label)
        sheet.cell(row=row_index, column=value_start, value=value)

    for column_letter in ["A", "B", "C", "D", "E", "F", "G", "H"]:
        sheet.column_dimensions[column_letter].width = 12
    for row_index in [5, 10, 14, 15, 16, 17, 18]:
        sheet.row_dimensions[row_index].height = 22
    sheet.row_dimensions[6].height = 24
    sheet.row_dimensions[7].height = 24
    sheet.row_dimensions[9].height = 22
    sheet.row_dimensions[11].height = 28
    sheet.row_dimensions[12].height = 8
    sheet.freeze_panes = None
    apply_common_sheet_settings(sheet, "$1:$4", "A1:H18")


def create_workbook(payload):
    styles = build_styles()
    workbook = Workbook()
    workbook._named_styles["Normal"].font = Font(
        name=get_report_font_name(),
        size=10,
        color=TEXT_NAVY
    )
    summary_sheet = workbook.active
    summary_sheet.title = "SUMMARY"
    write_summary_sheet(summary_sheet, payload, styles)

    review_groups_sheet = workbook.create_sheet("Review Groups")
    apply_table(
        review_groups_sheet,
        "Review Group Summary",
        [
            "Category",
            "Item Type",
            "QC Item",
            "Severity",
            "Count",
            "Sample Items",
            "Recommendation"
        ],
        payload.get("review_groups", []),
        styles,
        4,
        [18, 20, 24, 12, 10, 45, 45],
        count_column=5,
        numeric_columns=[5],
        center_columns=[4]
    )

    key_samples_sheet = workbook.create_sheet("Key Samples")
    apply_table(
        key_samples_sheet,
        "Key Review Samples",
        [
            "Category",
            "Item Type",
            "Item Name",
            "Severity",
            "QC Item",
            "Review Message",
            "Recommendation"
        ],
        payload.get("key_samples", []),
        styles,
        4,
        [18, 20, 28, 12, 22, 45, 45],
        center_columns=[4]
    )

    full_detail_sheet = workbook.create_sheet("Full Detail")
    apply_table(
        full_detail_sheet,
        "DOC QC Detail",
        [
            "Category",
            "Severity",
            "ElementId",
            "ElementType",
            "Name",
            "QC Item",
            "Review Message",
            "Current Value",
            "Recommendation"
        ],
        payload.get("full_detail", []),
        styles,
        2,
        [16, 11, 12, 18, 28, 20, 42, 30, 42],
        center_columns=[2, 3],
        detail_mode=True
    )

    return workbook


def main(arguments):
    if len(arguments) != 3:
        print(
            "Usage: make_styled_xlsx.py <input_json_path> <output_xlsx_path>",
            file=sys.stderr
        )
        return 2

    input_json_path = os.path.abspath(arguments[1])
    output_xlsx_path = os.path.abspath(arguments[2])

    try:
        payload = load_payload(input_json_path)
        workbook = create_workbook(payload)
        workbook.save(output_xlsx_path)
        print(output_xlsx_path)
        return 0
    except Exception as ex:
        print("Styled XLSX helper failed: {0}".format(ex), file=sys.stderr)
        return 1


if __name__ == "__main__":
    sys.exit(main(sys.argv))
